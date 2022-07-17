from ast import Return
from tkinter import *
from click import command

FONT_1 = ('Lucide Handwritting', 10, 'normal')
data_list = [['Word'], ["Meaning"], ["Phenotics"]]
len_list = 0

#_________ get word _____________3

def get_word():
    word = word_input.get()
    data_list = search_row(word.lower())
    if data_list[0]:
        update_ui('update', data=data_list)
    else:
        update_ui('update', [['NONE'], ["NOTHING FOUND"], ['EMPTY']])
    word_entered(w=word)
    update_ui('load')


def next_word():
    update_ui( 'next')
    update_ui('load')

def back_word():
    update_ui( 'back')
    update_ui('load')

#_________Find Word, return arrays_________#
import openpyxl as xl


ws = xl.load_workbook('H:\\Visual_Studio\\Pyton_project\\Dictnory\\Files\\Language.xlsx')
opt = ['Ālu Kuṟumba', 'Belari', 'Brahui', 'Gadba', 'Gondi', 'Iruḷa', 'Kannaḍa', 'Kota', 'Koḍagu', 'Kolami', 'Koraga', 'Kuṛux', 'Beṭṭa Kuruba', 'Malayalam', 'Malto', 'Manḍa', 'Naikṛi', 'Naiki of Chanda', 'Parji', 'Pālu Kuṟumba', 'proto-Dravidian', 'Pengo', 'Tamil', 'Telugu', 'Toda', 'Tulu', 'Konḍa', 'Kui', 'Kuwi']
opt_unicode = ['ಆಲು ಕುರುಂಬ','ಬೆಲರಿ','ಬ್ರರಹುಯ್','ಗಡ್ಬ', 'ಗೊಂಡಿ','ಇರುಳ', 'ಕನ್ನಡ','ಕೊತ','ಕೊಡವ','ಕೊಲಮಿ', 'ಕೊರಗ', 'ಕುರು', 'ಬೆಟ್ಟ ಕುರುಬ', 'ಮಲೆಯಾಳಂ', 'ಮಲ್ಟೊ','ಮಂಡ', 'ನೈಕ್ರಿ', 'ನೈಕ್ರಿ ಚಂದ', 'ಪರ್ಜಿ', 'ಪಾಲು ಕುರುಂಬ', 'ಅಡಿ ಪದ', 'ಪೆಂಗೊ', 'தமிழ்', 'తెలుగు','ತೊಡ' , 'ತುಳು', 'ಕೊಂಡ', 'ಕುಯ್', 'ಕುವೈ']
opt_ಬರಹ = ['ಕನ್ನಡ', 'తెలుగు', 'தமிழ்']
ಬರಹ = ['C', 'E', 'D']
meaning = 'F'

def search_row(word:str):
    if word == '':
        return [ ['Word'], ["Phenotics"], ["Meaning"]]
    p_list = []
    m_list = []
    x_list = []
    word_list = []
    english_word = []
    meaning_list = []
    page = opt[opt_unicode.index(language_selected.get())]
    sheet = ws[page]
    column = ಬರಹ[opt_ಬರಹ.index(ಬರಹ_selected.get())]
    for col in sheet.iter_cols(min_col=6, max_col=6):
        for cell in col:
            temp = str(cell.value)
            temp = temp.split(',')
            for x in temp:
                if x == word:
                    p_list.append(str(cell.row))
                x = x.split(' ')
                for i in x:
                    if i == word:
                        m_list.append(str(cell.row))
                    if word in i:
                        x_list.append(str(cell.row))

    for x in p_list:
        word_list.append(sheet[column + str(x)].value)
        english_word.append(sheet['G' + str(x)].value)
        meaning_list.append(sheet["F" + str(x)].value)

    for x in m_list:
        if x in p_list:
            continue
        word_list.append(sheet[column + str(x)].value)
        english_word.append(sheet['G' + str(x)].value)
        meaning_list.append(sheet["F" + str(x)].value)
    
    for x in x_list:
        if x in p_list or x in m_list:
            continue
        word_list.append(sheet[column + str(x)].value)
        english_word.append(sheet['G' + str(x)].value)
        meaning_list.append(sheet["F" + str(x)].value)
    
    return [english_word, meaning_list, word_list]

#_________windows__________3

w = Tk()
w.title('ಪದನರಕೆ')
w.config(padx=20, pady=20)

tell_search = Label(text="ಬೇಕಿರುವ ಪದ :", font=FONT_1)
tell_search.grid(row=1, column=0, pady=10)

ನುಡಿ_search = Label(text="ನುಡಿ :", font=FONT_1)
ನುಡಿ_search.grid(row=0, column=0)

ನುಡಿ_search1 = Label(text="ಬರಹದ ನುಡಿ :", font=FONT_1)
ನುಡಿ_search1.grid(row=0, column=3)

word_input = Entry(width=60, bd=5)
word_input.focus_set()
word_input.grid(row=1, column=1, columnspan= 2)

search_button = Button(text="ಹುಡಕು", width=12, command=get_word)
search_button.grid(row=2, column=1, columnspan = 2)
w.bind('<Return>', lambda event: get_word())

language_selected = StringVar()
language_selected.set('ಕನ್ನಡ')
language_option = OptionMenu(w, language_selected, *opt_unicode)
language_option.grid(row=0, column=1)

ಬರಹ_selected = StringVar()
ಬರಹ_selected.set('ಕನ್ನಡ')
ಬರಹ_option = OptionMenu(w, ಬರಹ_selected, *opt_ಬರಹ)
ಬರಹ_option.grid(row=0, column=4)

ನುಡಿ_out = Label(text="Word", font=FONT_1, width=20, height=10)
ನುಡಿ_out.grid(row=3, column=0)

entered_word = Label(text="", font=FONT_1)
entered_word.grid(row=1, column=4)

ನುಡಿ_out1 = Label(text="Meaning", font=FONT_1, width=80, height=10)
ನುಡಿ_out1.grid(row=3, column=1 , columnspan = 2)

ನುಡಿ_out2 = Label(text="Phonotic", font=FONT_1, width=20, height=10)
ನುಡಿ_out2.grid(row=3, column=3)

data_out1 = Label(text="No of Words", font=FONT_1, width=20, height=10)
data_out1.grid(row=2, column=4)

data_out2 = Label(text="word no", font=FONT_1, width=20, height=10)
data_out2.grid(row=2, column=0)

search_next = Button(text="ಮುಂದಿನ ಪದ", width=12, command=next_word)
search_next.grid(row=4, column=2)
w.bind('<Right>', lambda event: next_word())

search_back = Button(text="ಹಿಂದಿನ ಪದ", width=12, command=back_word)
search_back.grid(row=4, column=0)
w.bind('<Left>', lambda event: back_word())

#________ display finctions___________#


def update_ui(func:str, data = [["Phenotics"], ["Meaning"], ['Word']]):

    if update_ui.current >= update_ui.len:
        update_ui.current = 0
    
    if update_ui.current < 0:
        update_ui.current = update_ui.len - 1

    if func == 'load':
        ನುಡಿ_out['text'] = update_ui.dta[0][update_ui.current].replace(',', '\n')
        ನುಡಿ_out1['text'] = update_ui.dta[1][update_ui.current].replace(',', '\n')
        ನುಡಿ_out2['text'] = update_ui.dta[2][update_ui.current].replace(',', '\n')
        data_out2['text'] = 'ಇಗಿನ ಪದ ' + str(update_ui.current + 1)
    
    if func == 'update':
        if data[0] == '':
            update_ui.dta = [["Phenotics"], ["Meaning"], ['Word']]
        else:
            update_ui.dta = data
        update_ui.current = 0
        update_ui.len = len(data[0])
        data_out1['text'] = str(update_ui.len)
        # print(update_ui.len)
    
    if func == 'next':
        update_ui.current += 1
    
    if func == 'back':
        update_ui.current -= 1
    


update_ui.current = 0
update_ui.len = 0
update_ui.dta = [['Word'], ["Meaning"], ["Phenotics"]]


def word_entered(w:str):
    entered_word['text'] = w
    word_input.delete(0,END)
    word_input.focus_set()

w.mainloop()

